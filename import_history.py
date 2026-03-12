"""
Import historic padel match data from Excel into Supabase.
Run: python import_history.py
"""
import re
import sys
import requests
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = "https://kyhcstibhrmksipcgujc.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imt5aGNzdGliaHJta3NpcGNndWpjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMyODgxNTYsImV4cCI6MjA4ODg2NDE1Nn0.8p-0jgNKohfmOZcFHxCjvTE6lHpXcrPQBcBSYgWOLk8"
EXCEL_FILE   = r"c:\Users\wesl3\Desktop\Padel League 2026_Updated.xlsx"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# ── Name normalisation ────────────────────────────────────────────────────────
NAME_MAP = {
    "aron":   "Aaron",
    "aaron":  "Aaron",
    "boffa":  "Boffa",
    "cade":   "Cade",
    "case":   "Cade",   # typo in sheet
    "dan":    "Dan",
    "gareth": "Gareth",
    "josh":   "Josh",
    "kurt":   "Kurt",
    "paul":   "Paul",
    "steve":  "Steve",
    "terry":  "Terry",
    "wes":    "Wes",
    "stef":   "Stef",
}

def norm(name):
    return NAME_MAP.get(name.strip().lower(), name.strip().title())

# ── Match line parser ─────────────────────────────────────────────────────────
def parse_match(line):
    line = line.replace("\xa0", " ").replace("\u2060", "").strip()
    line = re.sub(r"\(.*?\)", "", line).strip()   # remove "(draw ...)" notes

    m = re.search(r"(\d+)\s*[–\-]\s*(\d+)", line)
    if not m:
        return None
    s1, s2 = int(m.group(1)), int(m.group(2))
    left  = line[:m.start()].strip()
    right = line[m.end():].strip()

    def split_team(s):
        s = s.strip().strip("&").strip()
        if "&" in s:
            parts = [p.strip() for p in re.split(r"\s*&\s*", s) if p.strip()]
        else:
            parts = [p.strip() for p in s.split() if p.strip()]
        return parts

    t1 = split_team(left)
    t2 = split_team(right)

    if len(t1) != 2 or len(t2) != 2:
        return None

    return norm(t1[0]), norm(t1[1]), s1, s2, norm(t2[0]), norm(t2[1])

# ── Supabase helpers ──────────────────────────────────────────────────────────
def sb_get(path, params=None):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS, params=params)
    r.raise_for_status()
    return r.json()

def sb_post(path, data):
    r = requests.post(f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS, json=data)
    r.raise_for_status()
    return r.json()

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    # List existing leagues
    leagues = sb_get("leagues", {"select": "id,name", "order": "created_at"})
    print("\nExisting leagues:")
    for i, l in enumerate(leagues):
        print(f"  [{i}] {l['name']}  (id: {l['id']})")
    print(f"  [n] Create new league")
    choice = input("\nEnter number or 'n': ").strip().lower()

    if choice == "n":
        name = input("New league name: ").strip()
        from nanoid import generate
        lid = generate(size=8)
        sb_post("leagues", {"id": lid, "name": name})
        league_id = lid
        print(f"Created league '{name}' with id {lid}")
    else:
        league_id = leagues[int(choice)]["id"]
        print(f"Using league id: {league_id}")

    # Load existing players
    existing_players = {p["name"].lower(): p["id"] for p in sb_get("players", {"league_id": f"eq.{league_id}", "select": "id,name"})}

    def get_or_create_player(name):
        key = name.lower()
        if key not in existing_players:
            result = sb_post("players", {"league_id": league_id, "name": name})
            pid = result[0]["id"]
            existing_players[key] = pid
            print(f"  + Player: {name}")
        return existing_players[key]

    # Load already-imported session dates
    existing_sessions = {s["date"] for s in sb_get("sessions", {"league_id": f"eq.{league_id}", "select": "date"})}

    # Process each session sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    session_sheets = [s for s in wb.sheetnames if re.match(r"\d{2}\.\d{2}\.\d{4}", s)]

    for sheet_name in session_sheets:
        # Convert DD.MM.YYYY → YYYY-MM-DD
        d, mo, y = sheet_name.split(".")
        date_iso = f"{y}-{mo}-{d}"
        label = f"Session – {int(d)} {['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][int(mo)]} {y}"

        ws = wb[sheet_name]
        all_rows = [row[0] for row in ws.iter_rows(values_only=True) if row[0] and isinstance(row[0], str)]

        # Find match lines (contain a score pattern)
        match_lines = [r for r in all_rows if re.search(r"\d+\s*[–\-]\s*\d+", r)]

        if date_iso in existing_sessions:
            print(f"\n-- {sheet_name}: already imported, skipping")
            continue

        if not match_lines:
            print(f"\n-- {sheet_name}: no match data — skipping")
            continue

        print(f"\n>> {sheet_name}: {len(match_lines)} matches")

        # Create session
        result = sb_post("sessions", {"league_id": league_id, "date": date_iso, "label": label})
        session_id = result[0]["id"]

        skipped = 0
        for line in match_lines:
            parsed = parse_match(line)
            if not parsed:
                print(f"  ⚠ Could not parse: {repr(line)}")
                skipped += 1
                continue

            p1n, p2n, s1, s2, p3n, p4n = parsed
            p1 = get_or_create_player(p1n)
            p2 = get_or_create_player(p2n)
            p3 = get_or_create_player(p3n)
            p4 = get_or_create_player(p4n)

            sb_post("matches", {
                "session_id": session_id,
                "scoring_type": "americano",
                "team1_p1": p1, "team1_p2": p2,
                "team2_p1": p3, "team2_p2": p4,
                "team1_score": s1, "team2_score": s2,
            })

        print(f"  OK: Imported {len(match_lines) - skipped} matches ({skipped} skipped)")

    print("\nImport complete!")

if __name__ == "__main__":
    main()
